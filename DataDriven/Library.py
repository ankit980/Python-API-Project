import requests
import json
import jsonpath
import openpyxl

class Common:
    def __init__(self,FileNamePath,SheetName):
        global  wk
        global sh
        wk=openpyxl.load_workbook(FileNamePath)
        sh=wk[SheetName]

    def fetch_row_count(self):
        rows=sh.max_row
        return rows

    def fetch_column_count(self):
        col=sh.max_column
        return col

    def fetch_key_names(self):
        c=sh.max_column
        l1=[]
        for i in range(1,c+1):
            cell=sh.cell(row=1,column=1)
            l1.insert(i-1,cell.value)
        return l1

    def update_request_with_data(self,rowNumber,jsonRequest,keyList):
        c=sh.max_column
        for i in range(1,c+1):
            cell=sh.cell(rowNumber,column=i)
            jsonRequest[keyList[i-1]]=cell.value
        return jsonRequest

